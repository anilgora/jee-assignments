#!/usr/bin/env python3
"""Build the difficulty-evaluation workbook from a ratings JSON file.

Usage:
    python build_report.py ratings.json output.xlsx

Input schema
------------
{
  "source_pdf": "Assignment_3.pdf",
  "evaluated_on": "2026-08-05",              # optional
  "questions": [
    {
      "sr_no": 1,                            # int, sequential from 1
      "q_ref": "Q1",                         # number as printed, e.g. "Q12(ii)"
      "page": 1,                             # int or null
      "question_type": "Single Correct Type", # exact string from taxonomy.md
      "chapter": "Definite Integration",
      "concept_1": "King's rule (a+b-x property)",
      "concept_2": "Periodic functions",     # "" if single-concept
      "scores": {"C": 4, "L": 3, "M": 2, "I": 4, "T": 3},   # each 1-5
      "est_time_min": 9,                     # number
      "difficulty_level": 3,                 # int 0-5, final call after overrides
      "key_step": "Recognise the period-4 collapse.",
      "justification": "Two to four sentences naming the concepts and the deciding step.",
      "confidence": "High"                   # High | Medium | Low
    }
  ]
}

Every field except "page" and "concept_2" is required. The script fails loudly on a
missing field rather than writing a half-built workbook.
"""

import json
import sys
from collections import Counter
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LEVEL_LABELS = {
    0: "Easier than JEE Main Level(Level 0)",
    1: "JEE Main Level(Level 1)",
    2: "A little easier than JEE Advanced(Level 2)",
    3: "JEE Advanced(Level 3)",
    4: "A little more difficult than JEE Advanced(Level 4)",
    5: "Very difficult than JEE Advanced(Level 5)",
}

LEVEL_FILLS = {
    0: "D9EAD3",  # light green
    1: "E8F0D8",
    2: "FFF2CC",  # amber
    3: "FCE5CD",
    4: "F4CCCC",  # light red
    5: "EA9999",  # deeper red
}

HEADER_FILL = PatternFill("solid", fgColor="1A365D")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=13, bold=True, color="1A365D")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EVAL_COLUMNS = [
    ("Sr No", 7),
    ("Q Ref", 9),
    ("Page", 6),
    ("Chapter", 22),
    ("Concept 1", 26),
    ("Concept 2", 24),
    ("Question Type", 24),
    ("Level", 7),
    ("Level Description", 34),
    ("C", 5),
    ("L", 5),
    ("M", 5),
    ("I", 5),
    ("T", 5),
    ("Composite", 11),
    ("Est. Time (min)", 13),
    ("Check", 22),
    ("Key Step", 40),
    ("Justification", 70),
    ("Confidence", 11),
]

BLUEPRINT_COLUMNS = [
    ("Sr No", 7),
    ("Chapter", 24),
    ("Concept1", 28),   # no space - matches the generation script
    ("Concept 2", 26),  # one space - matches the generation script
    ("Q Type", 26),
    ("Difficulty Level", 42),
]

REQUIRED = ["sr_no", "q_ref", "question_type", "chapter", "concept_1",
            "scores", "est_time_min", "difficulty_level", "key_step",
            "justification", "confidence"]


def fail(msg):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def validate(questions):
    for i, q in enumerate(questions):
        where = f"questions[{i}] (sr_no={q.get('sr_no', '?')})"
        for field in REQUIRED:
            if field not in q or q[field] in (None, ""):
                fail(f"{where}: missing required field '{field}'")
        lvl = q["difficulty_level"]
        if not isinstance(lvl, int) or lvl not in LEVEL_LABELS:
            fail(f"{where}: difficulty_level must be an integer 0-5, got {lvl!r}")
        for dim in "CLMIT":
            v = q["scores"].get(dim)
            if not isinstance(v, (int, float)) or not 1 <= v <= 5:
                fail(f"{where}: scores['{dim}'] must be between 1 and 5, got {v!r}")
        if q["confidence"] not in ("High", "Medium", "Low"):
            fail(f"{where}: confidence must be High, Medium or Low, "
                 f"got {q['confidence']!r}")


def composite_of(scores):
    return round(sum(scores[d] for d in "CLMIT") / 5, 1)


def implied_level(composite):
    if composite <= 1.5:
        return 0
    if composite <= 2.2:
        return 1
    if composite <= 2.9:
        return 2
    if composite <= 3.7:
        return 3
    if composite <= 4.4:
        return 4
    return 5


def style_header(ws, row_idx, columns):
    for col_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row_idx].height = 30


def build_evaluation(wb, data):
    ws = wb.active
    ws.title = "Evaluation"

    ws["A1"] = f"Difficulty Evaluation - {data.get('source_pdf', 'assignment')}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Evaluated on {data.get('evaluated_on', date.today().isoformat())} - "
                "scores C/L/M/I/T are conceptual depth, concept linkage, manipulation "
                "load, insight required, traps and precision, each 1-5. "
                "Level is the final call after the rubric override rules, which is why "
                "it can differ from the composite.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    ws.row_dimensions[2].height = 30

    style_header(ws, 4, EVAL_COLUMNS)

    for offset, q in enumerate(data["questions"]):
        r = 5 + offset
        s = q["scores"]
        comp = composite_of(s)
        lvl = q["difficulty_level"]
        gap = abs(lvl - implied_level(comp))
        check = "OK" if gap <= 1 else f"REVIEW: scores imply L{implied_level(comp)}"

        values = [
            q["sr_no"], q["q_ref"], q.get("page", ""), q["chapter"],
            q["concept_1"], q.get("concept_2", ""), q["question_type"],
            lvl, LEVEL_LABELS[lvl],
            s["C"], s["L"], s["M"], s["I"], s["T"],
            comp, q["est_time_min"], check,
            q["key_step"], q["justification"], q["confidence"],
        ]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=c in (4, 5, 6, 7, 9, 17, 18, 19))

        for col in (8, 9):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LEVEL_FILLS[lvl])
        ws.cell(row=r, column=8).font = Font(name="Arial", size=10, bold=True)

        if gap > 1:
            ws.cell(row=r, column=17).font = Font(name="Arial", size=10, bold=True,
                                                  color="CC0000")
        if q["confidence"] != "High":
            ws.cell(row=r, column=20).font = Font(name="Arial", size=10, bold=True,
                                                 color="CC6600")
        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:T{4 + len(data['questions'])}"
    return ws


def build_summary(wb, data, n_rows):
    ws = wb.create_sheet("Summary")
    last = 4 + n_rows

    ws["A1"] = "Summary"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Level distribution"
    ws["A3"].font = Font(name="Arial", size=11, bold=True)
    style_header(ws, 4, [("Level", 8), ("Description", 44), ("Count", 9),
                         ("Share", 10)])

    for lvl in range(6):
        r = 5 + lvl
        ws.cell(row=r, column=1, value=lvl).font = BODY_FONT
        ws.cell(row=r, column=2, value=LEVEL_LABELS[lvl]).font = BODY_FONT
        c = ws.cell(row=r, column=3,
                    value=f"=COUNTIF(Evaluation!$H$5:$H${last},A{r})")
        c.font = BODY_FONT
        p = ws.cell(row=r, column=4, value=f"=IF($C$11=0,0,C{r}/$C$11)")
        p.font = BODY_FONT
        p.number_format = "0.0%"
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
            ws.cell(row=r, column=col).fill = PatternFill("solid",
                                                          fgColor=LEVEL_FILLS[lvl])

    ws.cell(row=11, column=2, value="Total").font = Font(name="Arial", size=10, bold=True)
    tot = ws.cell(row=11, column=3, value="=SUM(C5:C10)")
    tot.font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=12, column=2, value="Mean level").font = BODY_FONT
    mean = ws.cell(row=12, column=3, value=f"=AVERAGE(Evaluation!$H$5:$H${last})")
    mean.font = BODY_FONT
    mean.number_format = "0.00"
    ws.cell(row=13, column=2, value="Total estimated solve time (min)").font = BODY_FONT
    ws.cell(row=13, column=3, value=f"=SUM(Evaluation!$P$5:$P${last})").font = BODY_FONT

    chapters = sorted({q["chapter"] for q in data["questions"]})
    start = 16
    ws.cell(row=start - 1, column=1, value="Chapter distribution").font = Font(
        name="Arial", size=11, bold=True)
    style_header(ws, start, [("Chapter", 30), ("Count", 9), ("Mean level", 12)])
    for i, ch in enumerate(chapters):
        r = start + 1 + i
        ws.cell(row=r, column=1, value=ch).font = BODY_FONT
        ws.cell(row=r, column=2,
                value=f'=COUNTIF(Evaluation!$D$5:$D${last},A{r})').font = BODY_FONT
        m = ws.cell(row=r, column=3,
                    value=f'=IF(B{r}=0,"",AVERAGEIF(Evaluation!$D$5:$D${last},'
                          f'A{r},Evaluation!$H$5:$H${last}))')
        m.font = BODY_FONT
        m.number_format = "0.00"
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER

    ws.cell(row=start + 3 + len(chapters), column=1,
            value="Counts are live formulas over the Evaluation sheet, so edits "
                  "to a level there flow through here.").font = Font(
        name="Arial", size=9, italic=True, color="555555")
    return ws


def build_blueprint(wb, data):
    ws = wb.create_sheet("Blueprint")
    ws["A1"] = "Blueprint - column layout matches Question_Paper_Blueprint.xlsx"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Copy this sheet into a standalone .xlsx to feed the generation script. "
                "Header spelling, including 'Concept1' without a space, must not be "
                "changed.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    style_header(ws, 4, BLUEPRINT_COLUMNS)
    for offset, q in enumerate(data["questions"]):
        r = 5 + offset
        values = [q["sr_no"], q["chapter"], q["concept_1"],
                  q.get("concept_2", ""), q["question_type"],
                  LEVEL_LABELS[q["difficulty_level"]]]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A5"
    return ws


def main():
    if len(sys.argv) != 3:
        fail("usage: python build_report.py ratings.json output.xlsx")

    with open(sys.argv[1], encoding="utf-8") as fh:
        data = json.load(fh)

    questions = data.get("questions")
    if not questions:
        fail("no questions found in the input JSON")
    validate(questions)

    wb = Workbook()
    build_evaluation(wb, data)
    build_summary(wb, data, len(questions))
    build_blueprint(wb, data)
    wb.save(sys.argv[2])

    dist = Counter(q["difficulty_level"] for q in questions)
    flagged = [q["sr_no"] for q in questions
               if abs(q["difficulty_level"]
                      - implied_level(composite_of(q["scores"]))) > 1]
    low_conf = [q["sr_no"] for q in questions if q["confidence"] != "High"]

    print(f"Wrote {sys.argv[2]} - {len(questions)} questions")
    print("Level distribution: "
          + ", ".join(f"L{l}={dist.get(l, 0)}" for l in range(6)))
    if flagged:
        print(f"Rows flagged for review (level vs scores): {flagged}")
    if low_conf:
        print(f"Rows below High confidence: {low_conf}")
    print("Reminder: run recalc on the file so the Summary formulas hold values.")


if __name__ == "__main__":
    main()
